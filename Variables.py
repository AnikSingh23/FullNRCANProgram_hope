import os
from operator import truediv

import pandas as pd
import numpy as np
import sys
import time
import xlrd
import openpyxl
# from xls2xlsx import XLS2XLSX as xlsx # for conversion process 1 I don't think is necessary anymore since either 2 or
# 4 are the preferred methods
from win32com.client import Dispatch
from fancyimpute import IterativeImputer
from fancyimpute import SoftImpute
import xgboost as xg
import argparse
from bs4 import BeautifulSoup as bs
import requests
from sklearn.model_selection import GridSearchCV
import matplotlib.pyplot as plt

extra_input = False
year_name = True

# Defining source folder to swap between just comment or uncomment the respective source folder line do not uncomment
# both
# source_folder = os.path.dirname(sys.executable)  # this line is for when using py installer so the path is correct
# for where the exe file resides
# source_folder = os.path.dirname(__file__)  # this line is for when running in pycharm will load the files into the
# project dir
source_folder = r'C:\Users\anik1\Desktop\Work\LEAP\leap-canada all scenarios_sperry et al._2023-03-16 - Copy (2)'  # for a specific folder

# setting up a piece of code which will change the input depending on the exe called will be called later at the
# input section. So what this does after compiling the program into an exe if running the cmd prompt you can call
# complete_program.exe -I to prompt for inputs OR you can create a shortcut to the exe right-click the shortcut and in
# the target line after the quotations add a -I and the shortcut will prompt you for inputs
parser = argparse.ArgumentParser()
parser.add_argument("-I", action="store_true", help="Trigger for extra input")
parser.add_argument("-i", action="store_true", help="Trigger for extra input")

parser.add_argument("-C", action="store_true", help="Trigger for conversion process")
parser.add_argument("-c", action="store_true", help="Trigger for conversion process")

parser.add_argument("-R", action="store_true", help="Trigger for final name conversion style 1")
parser.add_argument("-r", action="store_true", help="Trigger for final name conversion style 1")

parser.add_argument("-T", action="store_true", help="Trigger for final name conversion style 2")
parser.add_argument("-t", action="store_true", help="Trigger for final name conversion style 2")

args = parser.parse_args()

# Create the path for the temp_folder, so it is easier to call
temp_folder = source_folder + "\\temp\\"

# Creating a list of each province/folder name
Province = ["AB", "BC", "ATL", "CAN", "MB", "ON", "QC", "NB", "NL", "PE", "NS", "SK", "TER", "Temp"]

# Creates folders for each of the items in the Province list
for x in Province:
    mypath = source_folder + "\\" + x  # Creates paths to expected folders
    if not os.path.isdir(mypath):  # Checks if path exists if not create the path
        os.makedirs(mypath)  # Creates path if it does not exist


# Defining a simple function for grabbing href
def get_soup(url):  # defines the get_soup function for later use
    return bs(requests.get(url).text, 'html.parser')


# This will check the residential Canada page to see if it has updated (this assumes all the pages update to the same year at the same time which I believe is safe to assume)
update_check_url = "https://oee.nrcan.gc.ca/corporate/statistics/neud/dpa/menus/trends/comprehensive/trends_res_ca.cfm"
# Gets the href of the download which should include the year of publication and splits it using the "/" character where the split occurs (This assumes the format remains the same on the NRCAN side with all the files being put into a folder with the year attached directly)
update_link = get_soup(update_check_url).find(title='Click here to download all of the tables in this menu').get("href").split('/')
# Grabs the second last item from the split list which should always be the year and sets the variable for later
YearUpdated = "_" + update_link[-2]

# Note YearUpdated[1:] is used to remove the leading _ from this statement
print("The files are from " + YearUpdated[1:] + " if user input is not enabled, files will be named with this year")
print()

# If the extra argument I is added this code adds the extra input
if args.I or args.i or extra_input:
    print("------------------------------------------------------------------------------------------------------------------------")
    # This file is so the year can be input first then the scraper runs in the larger exe file so the exe can be started
    # and the user can walk away if necessary

    # This section should be clear it is just defining user input to append onto the eventual extracted file
    print('Enter text/year to add to the end of the filename generally the year. Will add it in by inserting an _ between')
    print('filename and input (has to follow normal windows file name conventions so no /\:*?"<>| symbols) Will overwrite')
    Year = input('files of the same name). Type * to use the default setting. If nothing is to be added just press enter: ')
    Year.strip()
    if Year == '':
        YearInput = ""
    elif Year == "*":
        YearInput = YearUpdated
    else:
        YearInput = "_" + Year

    print()

    print('Enter text/year to add to the end of the old filenames (if any exist) generally the year.')
    print('This has to follow normal windows file name conventions so no /\:*?"<>| symbols. Type * to use the default setting.')
    OldYear = input('If nothing is to be added just press enter: ')
    OldYear.strip()
    if OldYear == '':
        OldYearInput = ""
    elif OldYear == "*":
        OldYearInput = "Undefined"
    else:
        OldYearInput = " " + OldYear

    print()

    # A function to determine if the user wants to overwrite the data
    print("If .xlsx files already exist should the program overwrite them. Note if the program does not overwrite the "
          "base files and the file is already imputed nothing will occur due to the data already existing in a completed "
          "state.")
    overwrite = input('If yes then type Y/y if no then type N/n anything else will repeat the prompt: ')
    # Create an exclusion list so the user is less likely to hit a random button and trigger the program
    exclusion = [str("Y"), str("y"), str("N"), str("n")]
    # Repeats the user input until it gets the appropriate result
    while overwrite not in exclusion:
        overwrite = input("Please enter Y/y to overwrite files or N/n to not overwrite files anything else will repeat the prompt: ")
    # overwrite = "y"

    print()
    print()

    # Setting up a command to trigger levels of information specifically the time each process takes (mainly useful for debugging)
    print("Input for level of verbosity (at this point it is just a trigger for displaying section process times)")
    verbosity = input('Display extra information Y/y or N/n anything else will repeat the prompt: ')
    # Create an exclusion list so the user is less likely to hit a random button and trigger the program
    exclusion1 = [str("Y"), str("y"), str("N"), str("n")]
    # Repeats the user input until it gets the appropriate result
    while verbosity not in exclusion1:
        verbosity = input("Please enter Y/y to display extra info or N/n to not display extra info anything else will repeat the prompt: ")
    # verbosity = "y"

    print("------------------------------------------------------------------------------------------------------------------------")
    print()
else:  # Setting up the default settings for the inputs in the case extra input is not needed
    YearInput = YearUpdated
    OldYearInput = "Undefined"
    overwrite = "Y"
    verbosity = "N"

# Variable controlling whether input is asked for at the end of the program or not (semi depreciated this was for individual imputers)
end_variable = True

# Iterative imputer (base without changing estimator)
# imputer = Fimput(max_iter=100, keep_empty_features=True)

# Set verbosity level
verbose_level = 2 if verbosity.lower() == 'y' else 0

# Define the estimator with random_state=None to ensure randomness
estimator = xg.XGBRegressor(random_state=None)

# Initialize the IterativeImputer with dynamic verbosity and randomness
imputer = IterativeImputer(
    estimator=estimator,
    initial_strategy='median',
    imputation_order='random',
    min_value=0,
    max_iter=1000,
    skip_complete=True,
    verbose=verbose_level,
    random_state=None
)

# Backup imputer in case the main imputer throws an error
baImputer = IterativeImputer(
    estimator=estimator,
    initial_strategy='median',
    imputation_order='random',
    min_value=0,
    max_iter=1000,
    skip_complete=True,
    verbose=verbose_level,
    random_state=None
)
# Defining an empty list to count the missing values
misslist = []


def save_imputed_data_plot(dfC1_imputedtrans, Filename1):
    # Create a plot
    plt.figure(figsize=(10, 6))

    # Convert data to numeric where possible, coerce errors to NaN
    df_numeric = dfC1_imputedtrans.apply(pd.to_numeric, errors='coerce')

    # Plot each row as a separate line
    for idx in df_numeric.index:
        plt.plot(
            df_numeric.columns,
            df_numeric.loc[idx, :],
            label=f'Row {idx}'
        )

    # Add labels and title
    plt.xlabel('Columns')
    plt.ylabel('Values')
    plt.title('Imputed Data: Each Row as a Separate Line')

    # Add legend to identify the lines
    plt.legend(loc='best', fontsize='small', ncol=2)

    # Save the plot to a file
    plt.savefig(Filename1, format='png', bbox_inches='tight')

    # Close the plot to free memory
    plt.close()


# Defining the functions to complete the imputation this does a bunch of processes from checking if values missing
# values exist, preserving some features which are not to be imputed, transposing and un transposing the data frames
# so the imputation reads the variables correctly, and informing the user of how many values are being worked.
def onedfImp(Col1, Row1, Col2, Row2, Filename1, Sheetname1):
    # Defining variables and math to be used later in the program
    # Defining Column Range
    Cols1 = Col1 + ":" + Col2
    # Defining Number of rows
    Rows1 = Row2 - Row1 + 1
    # Defining first set of start rows
    SRow1 = Row1 - 2
    # Defining the start column as a number by converting the Col1 and Col3 variables to numbers note the .upper()
    # ensures the the value is a capital letter to ensure - 64 gives the correct value if not it will throw an error
    # as lower case letters require a - 96 to give correct value
    # The if statement checks if it is a one variable or two (col A vs col AA and converts it to the proper col number)
    if len(Col1) == 1:
        Colnum1 = ord(Col1.upper()) - 64
    else:
        Colnum1 = ((ord(Col1[:1].upper()) - 64) * 26) + (ord(Col1[1:].upper()) - 64)

    # IF data changes this section in each for loop must also change for example the current area of interest is
    # row C13:V13 which is 1 row and cols C:V for totals and C15:V24 which is 10 rows and cols C:V for individual
    # energy use.
    # To change the cols (which equates to years) change the value of usecols = C:V to the column
    # needed in Excel. (Now automatically updated if using the proper variable from the column_check.py)
    # To change the rows the starting point is indicated by skiprows(0, x) x is 2 minus the row
    # number from excel so cell C13 starts at skiprows(0,11) this is possibly due to the inclusion of a header
    # tests with header = FALSE might render this part unnecessary. To set the number of rows simply just change
    # the nrows variable to the number of rows being considered.

    TempFilename1 = temp_folder + Filename1

    df0 = pd.read_excel(TempFilename1, na_values=['x', "X"], skiprows=range(0, SRow1), nrows=Rows1, usecols=Cols1, sheet_name=Sheetname1)

    # Timing Function for testing purposes

    LocalStartTime = time.time()

    # If function to save time by skipping data frames without a null value
    if df0.isnull().values.any():
        # Adds the sum of the missing values to a list which can be counted later
        misslist.append(df0.isnull().sum().sum())

        # Tell the user that there are missing values to be imputed in this data frame
        print("There are " + str(df0.isnull().sum().sum()) + " missing values to be imputed in cells " + Col1 + str(Row1) + ":" + Col2 + str(Row2))

        # This section is to replace sections of long text which should be left as is because I think X is missing data,
        # I would think n.a. means is the electricity type is not used at all, and - is the same just formatted
        # differently I don't know why
        df0 = df0.replace(['n.a.'], '0.0009000009')
        df0 = df0.replace(['–'], '0.0008000008')

        # Transpose x and y-axis because of the ways imputation algorithms read data
        dfTransC1 = df0.transpose()

        # Sets up names for each column so when a column gets dropped we know which one is dropped and where to place it
        # the names follow the structure of col1, col2, col3 ... etc
        column_names = ["col" + str(i) for i in range(dfTransC1.shape[1])]
        dfTransC1.columns = column_names

        # Performs the imputation with the follow code
        try:
            imputedC1 = imputer.fit_transform(dfTransC1)
        except (Exception,):
            imputedC1 = baImputer.fit_transform(dfTransC1)

        # Turn the imputed data back into a data frame with column names (see twodfimp for more details)
        dfC1_imputed = pd.DataFrame(imputedC1, columns=dfTransC1.dropna(axis=1, how='all').columns)

        # find missing columns by the original set of column names and names after the drop occurs
        missing_cols = list(set(dfTransC1.columns) - set(dfC1_imputed.columns))
        # for each missing column reinsert a row of 0s by stripping the col off of col1, col2, col3 etc. of the missing
        # columns adding one to the number (since it inserts before the column position) to get the value of the next
        # column and inserts a new column of appropriate name in an appropriate place
        for colu in missing_cols:
            colu1 = colu[3:]
            colu1 = int(colu1) + 1
            colu2 = "col" + str(colu1)
            col_pos = dfC1_imputed.columns.get_loc(colu2)
            dfC1_imputed.insert(col_pos, colu, 0)

        # Transpose the imputed data BACK to the original orientation
        dfC1_imputedtrans = dfC1_imputed.transpose()
        # Replace the specified changes back to strings (needed to be numbers so the imputer would run)
        dfC1_imputedtrans = dfC1_imputedtrans.replace([0.0009000009], 'N.A.')
        dfC1_imputedtrans = dfC1_imputedtrans.replace([0.0008000008], '–')
        df0_imputedtrans = dfC1_imputedtrans

        save_imputed_data_plot(dfC1_imputedtrans, Filename1)

        # Append DataFrame to existing Excel file
        with pd.ExcelWriter(TempFilename1, mode='a', if_sheet_exists='overlay') as writer:
            df0_imputedtrans.to_excel(writer, sheet_name=Sheetname1, startrow=SRow1 + 1, startcol=Colnum1 - 1, index=False, header=False)
    # setting up the timing function to be linked to the verbosity user input
    if verbosity == "y" or verbosity == "Y":
        # Timing function for testing purposes
        LocalEndTime = time.time()
        # Determine the time and convert to minutes and seconds
        LocalTimeMin, LocalTimeSec = divmod((LocalEndTime - LocalStartTime) / 60, 1.0)
        print("Section completion time: " + str(round(LocalTimeMin)) + " Minutes and " + str(round(LocalTimeSec * 60)) + " Seconds")


def twodfImp(Col1, Row1, Col2, Row2, Col3, Row3, Col4, Row4, Filename1, Sheetname1):
    # Defining variables and math to be used later in the program
    # Defining Column Range
    Cols1 = Col1 + ":" + Col2
    Cols2 = Col3 + ":" + Col4
    # Defining Number of rows
    Rows1 = Row2 - Row1 + 1
    Rows2 = Row4 - Row3 + 1
    # Defining first set of start rows
    SRow1 = Row1 - 2
    SRow2 = Row3 - 2

    # Defining the start column as a number by converting column letters to numbers
    def col_to_num(col):
        num = 0
        for c in col.upper():
            num = num * 26 + (ord(c) - ord('A') + 1)
        return num

    Colnum1 = col_to_num(Col1)
    Colnum2 = col_to_num(Col3)

    TempFilename1 = temp_folder + Filename1

    df0 = pd.read_excel(
        TempFilename1,
        na_values=['x', 'X'],
        skiprows=range(0, SRow1),
        nrows=Rows1,
        usecols=Cols1,
        sheet_name=Sheetname1
    )
    df1 = pd.read_excel(
        TempFilename1,
        na_values=['x', 'X'],
        skiprows=range(0, SRow2),
        nrows=Rows2,
        usecols=Cols2,
        sheet_name=Sheetname1
    )

    # Timing Function for testing purposes
    LocalStartTime = time.time()

    # If function to save time by skipping data frames without a null value
    if df0.isnull().values.any() or df1.isnull().values.any():
        # Count the missing values
        msum = df0.isnull().sum().sum() + df1.isnull().sum().sum()

        # Adds the sum of the missing values to a list which can be counted later
        misslist.append(msum)

        # Tell the user that there are missing values to be imputed in this data frame
        print(
            "There are "
            + str(msum)
            + " missing values to be imputed in cells "
            + Col1
            + str(Row1)
            + ":"
            + Col2
            + str(Row2)
            + " and "
            + Col3
            + str(Row3)
            + ":"
            + Col4
            + str(Row4)
        )

        # Replace 'n.a.', 'N.A.', and '–' with temporary numeric values for imputation
        df0 = df0.replace(['n.a.', 'N.A.', '–'], [0.0009000009, 0.0009000009, 0.0008000008])
        df1 = df1.replace(['n.a.', 'N.A.', '–'], [0.0009000009, 0.0009000009, 0.0008000008])

        # Combine data frames
        dfcombine1 = pd.concat([df0, df1], ignore_index=True)

        # Transpose x and y-axis because of the ways imputers read data
        dfTransC1 = dfcombine1.transpose()

        # Set up column names
        column_names = ["col" + str(i) for i in range(dfTransC1.shape[1])]
        dfTransC1.columns = column_names

        # Initialize iteration variables
        max_iterations = 5
        iteration = 0
        acceptable = False

        while not acceptable and iteration < max_iterations:
            iteration += 1

            # Perform imputation
            try:
                imputedC1 = imputer.fit_transform(dfTransC1)
            except Exception:
                imputedC1 = baImputer.fit_transform(dfTransC1)

            # Turn the imputed data back into a data frame
            dfC1_imputed = pd.DataFrame(
                imputedC1,
                columns=dfTransC1.dropna(axis=1, how='all').columns
            )

            # Handle missing columns
            missing_cols = list(set(dfTransC1.columns) - set(dfC1_imputed.columns))
            missing_cols.sort(reverse=True)
            for colu in missing_cols:
                colu0 = int(colu[3:])

                # Column one ahead
                colu1 = colu0 + 1
                colu2 = "col" + str(colu1)

                # Column one behind
                colu3 = colu0 - 1
                colu4 = "col" + str(colu3)

                try:
                    col_pos = dfC1_imputed.columns.get_loc(colu2)
                except KeyError:
                    col_pos = dfC1_imputed.columns.get_loc(colu4) + 1

                dfC1_imputed.insert(col_pos, colu, 0)

            # Transpose back
            dfC1_imputedtrans = dfC1_imputed.transpose()

            # Replace back the strings
            dfC1_imputedtrans = dfC1_imputedtrans.replace(0.0009000009, 'N.A.')
            dfC1_imputedtrans = dfC1_imputedtrans.replace(0.0008000008, '–')

            # Now check the sums
            acceptable = True
            for col in dfC1_imputedtrans.columns:
                try:
                    print(dfC1_imputedtrans)
                    total_value = float(dfC1_imputedtrans.iloc[0][col])
                    component_values = dfC1_imputedtrans.iloc[1:, col].astype(float)
                    sum_components = component_values.sum()

                    if total_value != 0:
                        relative_difference = abs(sum_components - total_value) / abs(total_value)
                    else:
                        relative_difference = 0 if sum_components == 0 else float('inf')

                    if relative_difference > 0.10:
                        acceptable = False
                        print(
                            f"Iteration {iteration}: Total vs Sum in column '{col}' differs by more than 10% "
                            f"({relative_difference * 100:.2f}%). Re-imputing..."
                        )
                        break  # No need to check further columns

                except ValueError:
                    # Non-numeric values encountered, re-impute
                    acceptable = False
                    print(f"Iteration {iteration}: Non-numeric values encountered in column '{col}'. Re-imputing...")
                    break

            if not acceptable and iteration >= max_iterations:
                print(f"Warning: Unable to achieve acceptable totals after {max_iterations} iterations")

        # Plotting the imputed data
        save_imputed_data_plot(dfC1_imputedtrans, 'imputed_data_plot.png')

        # Break the combined imputed dataframe back into separate data frames
        df0_imputedtrans = dfC1_imputedtrans.iloc[:Rows1]
        df1_imputedtrans = dfC1_imputedtrans.iloc[Rows1:]

        # Append DataFrame to existing Excel file
        with pd.ExcelWriter(TempFilename1, mode='a', if_sheet_exists='overlay') as writer:
            df0_imputedtrans.to_excel(
                writer,
                sheet_name=Sheetname1,
                startrow=SRow1 + 1,
                startcol=Colnum1 - 1,
                index=False,
                header=False
            )
            df1_imputedtrans.to_excel(
                writer,
                sheet_name=Sheetname1,
                startrow=SRow2 + 1,
                startcol=Colnum2 - 1,
                index=False,
                header=False
            )

    # Timing function
    if verbosity.lower() == "y":
        LocalEndTime = time.time()
        total_time = LocalEndTime - LocalStartTime
        minutes = int(total_time // 60)
        seconds = int(total_time % 60)
        print(
            "Section completion time: "
            + str(minutes)
            + " Minutes and "
            + str(seconds)
            + " Seconds"
        )



def threedfImp(Col1, Row1, Col2, Row2, Col3, Row3, Col4, Row4, Col5, Row5, Col6, Row6, Filename1, Sheetname1):
    # Defining variables and math to be used later in the program
    # Defining Column Range
    Cols1 = Col1 + ":" + Col2
    Cols2 = Col3 + ":" + Col4
    Cols3 = Col5 + ":" + Col6

    # Defining Number of rows
    Rows1 = Row2 - Row1 + 1
    Rows2 = Row4 - Row3 + 1
    Rows3 = Row6 - Row5 + 1
    # Defining first set of start rows
    SRow1 = Row1 - 2
    SRow2 = Row3 - 2
    SRow3 = Row5 - 2

    # Defining the start column as a number
    def col_to_num(col):
        num = 0
        for c in col.upper():
            num = num * 26 + (ord(c) - ord('A') + 1)
        return num

    Colnum1 = col_to_num(Col1)
    Colnum2 = col_to_num(Col3)
    Colnum3 = col_to_num(Col5)

    TempFilename1 = temp_folder + Filename1

    df0 = pd.read_excel(
        TempFilename1,
        na_values=['x', 'X'],
        skiprows=range(0, SRow1),
        nrows=Rows1,
        usecols=Cols1,
        sheet_name=Sheetname1
    )
    df1 = pd.read_excel(
        TempFilename1,
        na_values=['x', 'X'],
        skiprows=range(0, SRow2),
        nrows=Rows2,
        usecols=Cols2,
        sheet_name=Sheetname1
    )
    df2 = pd.read_excel(
        TempFilename1,
        na_values=['x', 'X'],
        skiprows=range(0, SRow3),
        nrows=Rows3,
        usecols=Cols3,
        sheet_name=Sheetname1
    )

    # Timing Function for testing purposes
    LocalStartTime = time.time()

    # If function to save time by skipping data frames without a null value
    if df0.isnull().values.any() or df1.isnull().values.any() or df2.isnull().values.any():
        # Count the missing values
        msum = df0.isnull().sum().sum() + df1.isnull().sum().sum() + df2.isnull().sum().sum()

        # Adds the sum of the missing values to a list which can be counted later
        misslist.append(msum)

        # Tell the user that there are missing values to be imputed in this data frame
        print(
            "There are "
            + str(msum)
            + " missing values to be imputed in cells "
            + Col1
            + str(Row1)
            + ":"
            + Col2
            + str(Row2)
            + ", "
            + Col3
            + str(Row3)
            + ":"
            + Col4
            + str(Row4)
            + ", and "
            + Col5
            + str(Row5)
            + ":"
            + Col6
            + str(Row6)
        )

        # Replace 'n.a.', 'N.A.', and '–' with temporary numeric values for imputation
        df_list = [df0, df1, df2]
        for df in df_list:
            df.replace(['n.a.', 'N.A.', '–'], [0.0009000009, 0.0009000009, 0.0008000008], inplace=True)

        # Combine data frames
        dfcombine1 = pd.concat(df_list, ignore_index=True)

        # Transpose x and y-axis
        dfTransC1 = dfcombine1.transpose()

        # Set up column names
        column_names = ["col" + str(i) for i in range(dfTransC1.shape[1])]
        dfTransC1.columns = column_names

        # Initialize iteration variables
        max_iterations = 5
        iteration = 0
        acceptable = False

        while not acceptable and iteration < max_iterations:
            iteration += 1

            # Perform imputation
            try:
                imputedC1 = imputer.fit_transform(dfTransC1)
            except Exception:
                imputedC1 = baImputer.fit_transform(dfTransC1)

            # Turn the imputed data back into a data frame
            dfC1_imputed = pd.DataFrame(
                imputedC1,
                columns=dfTransC1.dropna(axis=1, how='all').columns
            )

            # Handle missing columns
            missing_cols = list(set(dfTransC1.columns) - set(dfC1_imputed.columns))
            missing_cols.sort(reverse=True)
            for colu in missing_cols:
                colu0 = int(colu[3:])

                # Column one ahead
                colu1 = colu0 + 1
                colu2 = "col" + str(colu1)

                # Column one behind
                colu3 = colu0 - 1
                colu4 = "col" + str(colu3)

                try:
                    col_pos = dfC1_imputed.columns.get_loc(colu2)
                except KeyError:
                    col_pos = dfC1_imputed.columns.get_loc(colu4) + 1

                dfC1_imputed.insert(col_pos, colu, 0)

            # Transpose back
            dfC1_imputedtrans = dfC1_imputed.transpose()

            # Replace back the strings
            dfC1_imputedtrans.replace(0.0009000009, 'N.A.', inplace=True)
            dfC1_imputedtrans.replace(0.0008000008, '–', inplace=True)

            # Now check the sums
            acceptable = True
            for col in dfC1_imputedtrans.columns:
                try:
                    total_value = float(dfC1_imputedtrans.iloc[0][col])
                    component_values = dfC1_imputedtrans.iloc[1:, col].astype(float)
                    sum_components = component_values.sum()

                    if total_value != 0:
                        relative_difference = abs(sum_components - total_value) / abs(total_value)
                    else:
                        relative_difference = 0 if sum_components == 0 else float('inf')

                    if relative_difference > 0.10:
                        acceptable = False
                        print(
                            f"Iteration {iteration}: Total vs Sum in column '{col}' differs by more than 10% "
                            f"({relative_difference * 100:.2f}%). Re-imputing..."
                        )
                        break

                except ValueError:
                    # Non-numeric values encountered
                    acceptable = False
                    print(f"Iteration {iteration}: Non-numeric values encountered in column '{col}'. Re-imputing...")
                    break

            if not acceptable and iteration >= max_iterations:
                print(f"Warning: Unable to achieve acceptable totals after {max_iterations} iterations")

        # Plotting the imputed data
        save_imputed_data_plot(dfC1_imputedtrans, 'imputed_data_plot.png')

        # Break the combined imputed dataframe back into separate data frames
        df0_imputedtrans = dfC1_imputedtrans.iloc[:Rows1]
        df1_imputedtrans = dfC1_imputedtrans.iloc[Rows1:Rows1 + Rows2]
        df2_imputedtrans = dfC1_imputedtrans.iloc[Rows1 + Rows2:]

        # Append DataFrame to existing Excel file
        with pd.ExcelWriter(TempFilename1, mode='a', if_sheet_exists='overlay') as writer:
            df0_imputedtrans.to_excel(
                writer,
                sheet_name=Sheetname1,
                startrow=SRow1 + 1,
                startcol=Colnum1 - 1,
                index=False,
                header=False
            )
            df1_imputedtrans.to_excel(
                writer,
                sheet_name=Sheetname1,
                startrow=SRow2 + 1,
                startcol=Colnum2 - 1,
                index=False,
                header=False
            )
            df2_imputedtrans.to_excel(
                writer,
                sheet_name=Sheetname1,
                startrow=SRow3 + 1,
                startcol=Colnum3 - 1,
                index=False,
                header=False
            )

    # Timing function
    if verbosity.lower() == "y":
        LocalEndTime = time.time()
        total_time = LocalEndTime - LocalStartTime
        minutes = int(total_time // 60)
        seconds = int(total_time % 60)
        print(
            "Section completion time: "
            + str(minutes)
            + " Minutes and "
            + str(seconds)
            + " Seconds"
        )





# Defining conversion method, methods 2 and 4 are the fastest versions. Method 2 cannot have Excel open but copies more
# formatting method 4 copies less formatting but Excel can be open.
def conversion(OriginalFileName, NewFileName):
    # Setting default conversion method
    method = 2
    # Checking the argument from earlier to determine the conversion method
    if args.C or args.c:
        method = 4

    if method == 1:
        # Longer conversion process which keeps the formatting (requires xls2xlsx and related dependencies)
        # (Roughly 50 min overall for every file)
        print("Conversion Method Temporarily Disabled")  # ConFile = xlsx(OriginalFileName)  # ConFile.to_xlsx(NewFileName)

    if method == 2:
        # Shorter Conversion process which keep formatting (requires pywin32 and a system with Excel installed)
        # Important to note I say roughly and round the time up because each run will have slightly different results
        # due to cpu load and memory load etc.

        # Possible issue with clicking away from the main application during a test when I clicked into pycharm the
        # process stalled on conversion, but I have also had a test where the entire time the program was not in focus, and
        # it ran without any trouble, so I am not sure

        # set up the dispatch of the Excel application
        xlApp = Dispatch('Excel.Application')
        # Create file paths
        oldName = temp_folder + OriginalFileName
        newName = temp_folder + NewFileName

        # opens the original file
        time.sleep(1)
        wb = xlApp.Workbooks.Open(oldName)
        # disables warnings for overwrite
        xlApp.DisplayAlerts = False
        time.sleep(1)
        # saves file as new name (note "51" is the file value for the format of .xlsx from the
        # https://learn.microsoft.com/en-us/office/vba/api/excel.xlfileformat page it is required otherwise it will throw
        # an error saying extension and file type are not matching)
        wb.SaveAs(newName, 51)
        time.sleep(1)
        # turn back on Excel based warnings
        xlApp.DisplayAlerts = True
        # close the workbook and exits excel (WARNING will close ALL instances of Excel open due to .Quit())
        wb.Close(True)
        time.sleep(0.5)
        xlApp.Application.Quit()

    if method == 3:
        # Short conversion process but does not keep formatting (Requires only pandas)
        # (roughly 20 min overall for every file)

        xls_file = pd.ExcelFile(OriginalFileName)
        sheet_names = xls_file.sheet_names

        # Create dict
        res = {}

        # Build dict of sheetname: dataframe of each sheet
        for sheet in sheet_names:
            res[sheet] = pd.read_excel(OriginalFileName, sheet_name=sheet, header=None)

        # Create ExcelWriter object
        with pd.ExcelWriter(NewFileName, engine="openpyxl") as writer:
            # Loop through dict, and have the writer write them to a single file
            for sheet, frame in res.items():
                frame.to_excel(writer, sheet_name=sheet, header=False, index=False)

    if method == 4:
        # Short conversion process keeping formatting (example use of chatGPT with some minor edits to ensure working condition)
        # (Roughly 24 min overall for every file)

        oldName = temp_folder + OriginalFileName
        newName = temp_folder + NewFileName

        # Open the .xls file using xlrd and preserving the formatting
        wb = xlrd.open_workbook(oldName, formatting_info=True)

        # Create a new workbook in openpyxl
        new_wb = openpyxl.Workbook()

        # Get the sheet names from the original xlrd workbook
        sheet_names = wb.sheet_names()

        # Loop through the sheet names and copy the sheets to the new workbook
        for sheet_name in sheet_names:
            # Get the sheet from the original workbook
            sheet = wb.sheet_by_name(sheet_name)

            # Create a new sheet in the new workbook
            new_sheet = new_wb.create_sheet(sheet_name)

            # Loop through the rows and columns in the original sheet and copy the cell values and formatting to the new sheet
            for row in range(sheet.nrows):
                for col in range(sheet.ncols):
                    new_cell = new_sheet.cell(row=row + 1, column=col + 1)
                    new_cell.value = sheet.cell_value(row, col)

                    # Set the cell formatting to match the original cell
                    xf_index = sheet.cell_xf_index(row, col)
                    xf = wb.xf_list[xf_index]
                    font = wb.font_list[xf.font_index]
                    new_font = openpyxl.styles.Font(name=font.name, size=10, bold=font.bold, italic=font.italic, strike=font.struck_out, )
                    new_cell.font = new_font
                    # new_cell.alignment = xf.alignment
                    # new_cell.border = xf.border
                    # new_cell.fill = xf.background
                    new_cell.number_format = wb.format_map[xf.format_key].format_str

            # Set the column widths in the new sheet to match the original sheet  # for col_idx, width in enumerate(sheet.col_widths):  #     new_sheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx + 1)].width = width

        del new_wb["Sheet"]

        # Save the new workbook as .xlsx
        new_wb.save(newName)

        # Open New workbook again and delete the original file


# Defining function for the final name change to be more similar to the leap structure, I made this a function, so it is
# easier to edit if need be plus it makes it easier to include the different arguments for the final name change
def LeapNameChange(CreatedFileName):
    # Checks if the arguments R or T are not enabled, if not then rename to be consistent with leap (note example name is tran_nl_e_imp.xlsx)
    if not args.R and not args.r and not args.T and not args.t and not year_name:
        temporary = os.path.basename(CreatedFileName)
        # Remove the .xlsx
        temporary = temporary[:-5]
        # Remove the _imp
        temporary = temporary[:-4]
        # Remove the User Input
        # UserCount = len(YearInput) * -1
        # temporary = temporary[:UserCount]
        # Remove the _e
        temporary = temporary[:-2]
        # Replace tr with ter for territories
        temporary = temporary.replace("_tr", "_ter")
        # Replace NF with NL for territories
        temporary = temporary.replace("_nf", "_nl")
        # Replace AGG with IND for aggregated industries (LEAP uses IND)
        temporary = temporary.replace("agg", "ind")
        # Replace Tran with Tra for transport (LEAP uses TRA)
        temporary = temporary.replace("tran", "tra")
        # Replace Tran with ca with can (LEAP uses Can)
        temporary = temporary.replace("ca", "can")
        if "BC_RES" not in temporary:
            temporary = temporary.replace("BCT", "BC")
        if "CAN_AGR" in temporary:
            temporary = temporary.replace("CAN_AGR", "AGR")
        # Upper case and replace _ with space
        temporary = temporary.upper().replace("_", " ")
        # Reverse Order of String
        s = temporary.split()[::-1]
        LeapNameList = []
        for i in s:
            # appending reversed words to list
            LeapNameList.append(i)
        # Join the LeapNameList using a space between the words by specifying a string before the .join
        LeapName = " ".join(LeapNameList)
        # Add back on .xlsx
        LeapName = LeapName + ".xlsx"
        # Turn back into the base name so the program can create the file structure as needed
        LeapName = os.path.basename(LeapName)
        # Allows the function to return a new value
        return LeapName

    if not args.R and not args.r and not args.T and not args.t and year_name:
        temporary = os.path.basename(CreatedFileName)
        # Remove the .xlsx
        temporary = temporary[:-5]
        # Remove the _imp
        temporary = temporary[:-4]
        # Find and keep the year input (skip the length of YearInput)
        UserCount = len(YearInput) * -1
        temp_year = temporary[UserCount:]
        temporary = temporary[:UserCount]
        # Now remove the _e (which happens after preserving the year input)
        temporary = temporary[:-2]
        # # Add the year input back to the string
        # temporary = temporary + temp_year
        # Replace tr with ter for territories
        temporary = temporary.replace("_tr", "_ter")
        # Replace NF with NL for territories
        temporary = temporary.replace("_nf", "_nl")
        # Replace AGG with IND for aggregated industries (LEAP uses IND)
        temporary = temporary.replace("agg", "ind")
        # Replace Tran with Tra for transport (LEAP uses TRA)
        temporary = temporary.replace("tran", "tra")
        # Replace Tran with ca with can (LEAP uses Can)
        temporary = temporary.replace("ca", "can")
        if "BC_RES" not in temporary:
            temporary = temporary.replace("BCT", "BC")
        if "CAN_AGR" in temporary:
            temporary = temporary.replace("CAN_AGR", "AGR")
        # Upper case and replace _ with space
        temporary = temporary.upper().replace("_", " ")
        # Reverse the order of words in the string
        s = temporary.split()[::-1]
        LeapNameList = []
        for i in s:
            # Appending reversed words to the list
            LeapNameList.append(i)
        # Join the LeapNameList using a space between the words
        LeapName = " ".join(LeapNameList)
        # Add back on .xlsx
        LeapName = LeapName + ".xlsx"
        # Return the newly created LeapName
        LeapName = os.path.basename(LeapName)
        return LeapName

    #  Check if the argument R is enabled if it is no changes to the file name need to be made
    if args.R or args.r:
        LeapName = CreatedFileName
        LeapName = os.path.basename(LeapName)
        return LeapName

    #  Check if the argument T is enabled if it is remove the _imp from the file names.
    if args.T or args.t:
        temporary = os.path.basename(CreatedFileName)
        # Remove the .xlsx
        temporary = temporary[:-5]
        # Remove the _imp
        temporary = temporary[:-4]
        # Add back in the .xlsx
        LeapName = temporary + ".xlsx"
        LeapName = os.path.basename(LeapName)
        return LeapName


# def checkvalues(filename, table, oldstyle):
#     # Examined_File = source_folder + filename
#
#     if oldstyle:
#         rows_skipped = 9
#     if not oldstyle:
#         rows_skipped = 10
#
#     # Read the file for the years column
#     dfcheck = pd.read_excel(filename, sheet_name=table, skiprows=rows_skipped, nrows=0)
#     # Turn column into a list
#     orig_year_list = dfcheck.columns.tolist()
#     # Create an alphabetical list the same size as the year list to show corresponding column letters
#     alphabetical_list = []
#     # This for loop will populate the alphabetical list with the letters if the number of columns surpass 26 (a-z) then this
#     # loop will add a preceding letter appropriately for example column 26 would be Z and 27 would be AA, column 52 would be
#     # AZ and column 53 would be BA
#     for i in range(len(orig_year_list)):
#         # Calculate the number of times the preceding character needs to be incremented
#         preceding_char_increments = i // 26
#         # Calculate the index of the current character in the alphabet (0-based)
#         char_index = i % 26
#         # Create the preceding characters by incrementing the character 'a' the number of times calculated
#         preceding_chars = ''.join([chr(97 + j) for j in range(preceding_char_increments)])
#         # Append the preceding characters and the current character to the result list
#         alphabetical_list.append(preceding_chars + chr(97 + char_index))
#
#     # Finds the min and max values of the columns excluding the first two (using 2:) since they are strings and are there
#     # for formatting the Excel table
#     print(orig_year_list)
#     orig_first_year = min(orig_year_list[2:])
#     orig_last_year = max(orig_year_list[2:])
#
#     # Finding the corresponding letter for the orig_first and orig_last year using the alphabetical list created earlier
#     orig_first_col = alphabetical_list[orig_year_list.index(orig_first_year)].upper()
#     orig_last_col = alphabetical_list[orig_year_list.index(orig_last_year)].upper()
#
#     print(orig_first_year, "   ", orig_first_col)
#     print(orig_last_year, "   ", orig_last_col)
#
#     # Return the first year, last year, and their corresponding Excel columns
#     return orig_first_year, orig_last_year, orig_first_col, orig_last_col

def checkvalues(filename, table, oldstyle):
    # Examined_File = source_folder + filename

    rows_skipped = 9 if oldstyle else 10

    # Read the file for the years column
    dfcheck = pd.read_excel(filename, sheet_name=table, skiprows=rows_skipped, nrows=0)

    # Turn column into a list
    orig_year_list = dfcheck.columns.tolist()

    # Get the first and last year (assuming the first two columns are not years)
    orig_first_year = min(orig_year_list[2:])
    orig_last_year = max(orig_year_list[2:])

    return orig_first_year, orig_last_year
